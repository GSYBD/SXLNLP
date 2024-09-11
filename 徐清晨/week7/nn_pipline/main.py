# -*- coding: utf-8 -*-

import torch
import os
import random
import os
import numpy as np
import logging
from config import Config
from model import TorchModel, choose_optimizer
from evaluate import Evaluator
from loader import load_data
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment

# [DEBUG, INFO, WARNING, ERROR, CRITICAL]
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

"""
模型训练主程序
"""

seed = Config["seed"]
random.seed(seed)
np.random.seed(seed)
torch.manual_seed(seed)
torch.cuda.manual_seed_all(seed)


def main(config):
    # 创建保存模型的目录
    if not os.path.isdir(config["model_path"]):
        os.mkdir(config["model_path"])
    # 加载训练数据
    train_data = load_data(config["train_data_path"], config)
    # print(len(train_data))
    # exit()
    # 加载模型
    model = TorchModel(config)
    # 标识是否使用gpu
    cuda_flag = torch.cuda.is_available()
    if cuda_flag:
        logger.info("gpu可以使用，迁移模型至gpu")
        model = model.cuda()
    # 加载优化器
    optimizer = choose_optimizer(config, model)
    # 加载效果测试类
    evaluator = Evaluator(config, model, logger)
    # 训练
    for epoch in range(config["epoch"]):
        epoch += 1
        model.train()
        # logger.info("epoch %d begin" % epoch)
        train_loss = []
        for index, batch_data in enumerate(train_data):
            if cuda_flag:
                batch_data = [d.cuda() for d in batch_data]

            optimizer.zero_grad()
            """
            train_data:一共被分成了12份，12份不知道哪里定义的
            
            input_ids, labels
             [ 101, 4659,  755,  ...,    0,    0,    0],
        [ 101,  517, 3457,  ...,    0,    0,    0],...
        一批是128样本，128句话，每句话最多30个字 128, 30
            """
            input_ids, labels = batch_data  # 输入变化时这里需要修改，比如多输入，多输出的情况
            # print(input_ids,labels)
            # print(input_ids.shape)
            # exit()
            # print(index)
            loss = model(input_ids, labels)
            loss.backward()
            optimizer.step()

            train_loss.append(loss.item())
        #     if index % int(len(train_data) / 2) == 0:
        #         logger.info("batch loss %f" % loss)
        # logger.info("epoch average loss: %f" % np.mean(train_loss))
        acc = evaluator.eval(epoch)
    # model_path = os.path.join(config["model_path"], "epoch_%d.pth" % epoch)
    # torch.save(model.state_dict(), model_path)  #保存模型权重
    return acc

def csv_split(path):
    df = pd.read_csv(path)
    total_rows = df.shape[0]

    # 分割csv文件成兩個，一個訓練集，一个验证集，样本各取一半
    split_index = 4000
    df1_part1 = df.iloc[:int(split_index/2)]
    df1_part2 = df.iloc[int(split_index/2):split_index]

    df2_part1 = df.iloc[split_index:split_index + int((total_rows - split_index) / 2)]
    df2_part2 = df.iloc[split_index + int((total_rows - split_index) / 2):]

    df_part1 = pd.concat([df1_part1, df2_part1])
    df_part2 = pd.concat([df1_part2, df2_part2])

    # 保存分割后的两个部分到新的 CSV 文件
    df_part1.to_csv('../data/train_cate_comment.csv', index=False)
    df_part2.to_csv('../data/valid_cate_comment.csv', index=False)
    # print(total_rows,df1_part1,df_part1,df_part2)

if __name__ == "__main__":
    # main(Config)
    # csv_split("../data/文本分类练习.csv")
    # for model in ["cnn"]:
    #     Config["model_type"] = model
    #     print("最后一轮准确率：", main(Config), "当前配置：", Config["model_type"])

    # 对比所有模型
    # 中间日志可以关掉，避免输出过多信息
    # 超参数的网格搜索
    # model_list = ["gated_cnn", "lstm", "gru", "bert", "bert_lstm"]
    model_list = ["gated_cnn", "lstm"]
    all_config = ['模型']
    model_excel = []
    for model in model_list:
        model_sheet = [model]
        Config["model_type"] = model
        for lr in [1e-3, 1e-4]:
            Config["learning_rate"] = lr
            for hidden_size in [128, 512]:
                Config["hidden_size"] = hidden_size
                for batch_size in [128]:
                    Config["batch_size"] = batch_size
                    for pooling_style in ["avg"]:
                        if model == model_list[0]:
                            all_config.append(
                                f'lr:{lr}\nhidden_size:{hidden_size}\nbatch_size:{batch_size}\npooling_style:{pooling_style}')

                        Config["pooling_style"] = pooling_style
                        acc = main(Config)
                        model_sheet.append(acc)
                        print("最后一轮准确率：", acc, "当前配置：", Config)
        model_excel.append(model_sheet)

    excel = [all_config] + model_excel

    # 存excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    for row_index, row_data in enumerate(excel, start=1):
        # print(row_index,row_data)
        for col_index, cell_data in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index, value=cell_data)
            sheet.cell(row=row_index, column=col_index).alignment = Alignment(wrapText=True)

    for i in range(len(all_config)):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = 20
    wb.save('./output/comment_model_comparison.xlsx')
