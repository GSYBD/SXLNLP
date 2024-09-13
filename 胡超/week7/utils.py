# -*- coding: utf-8 -*-
"""
author: Chris Hu
date: 2024/8/1
desc:
sample
"""

import math
import os
import pandas as pd
import random
import json
from week7.config import Config
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


def preprocess_data_and_config(csv_path):
    train_data_path = r'./data/train_data.json'
    test_data_path = r'./data/test_data.json'
    index_to_label = {0: "差评", 1: "好评"}
    label_to_index = {v: k for k, v in index_to_label.items()}

    data = pd.read_csv(csv_path, encoding='utf-8')
    new_columns = ['tag', 'title']
    data.columns = new_columns
    data['tag'] = data['tag'].replace(index_to_label)
    data = data.to_dict(orient='records')
    random.shuffle(data)
    train_data_len = math.floor(len(data) * .8)

    with open(train_data_path, 'w', encoding='utf-8') as f:
        for row in data[:train_data_len]:
            # below action will cause "JSON standard allows only one top-level value"
            # but in order not to change the original data loader, this is ok
            f.write(json.dumps(row, ensure_ascii=False) + '\r\n')

    with open(test_data_path, 'w', encoding='utf-8') as f:
        for row in data[train_data_len:]:
            f.write(json.dumps(row, ensure_ascii=False) + '\r\n')

    # update the Config in config.py
    Config['train_data_path'] = train_data_path
    Config['valid_data_path'] = test_data_path
    Config['label_to_index'] = label_to_index
    with open(r'./config.py', 'w', encoding='utf-8') as f:
        f.write('Config = ' + json.dumps(Config, ensure_ascii=False, indent=4))


def excel_report(results, excel_path):
    df = pd.DataFrame(results, index=[0]).drop(columns=['label_to_index', "model_path", "save_model"])
    sheet_name = "Sheet1"
    if os.path.exists(excel_path):
        # 读取现有的 Excel 文件
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # 将 DataFrame 写入 Excel 文件
            df.to_excel(writer, sheet_name=sheet_name, startrow=writer.sheets["Sheet1"].max_row, header=False,
                        index=False)
    else:
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        df.to_excel(excel_path, sheet_name=sheet_name, index=False)


def pretty_excel(excel_path):
    # 加载现有的工作簿
    wb = load_workbook(filename=excel_path)
    ws = wb.active

    def find_column_index(column_name):
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == column_name:
                    return get_column_letter(cell.column)
        return None

    acc_col_index = find_column_index('acc')
    # 设置字体样式
    font_style = Font(name='Arial', size=12, bold=True, italic=False, color='FFFFFF')
    # 设置背景颜色
    fill_color = PatternFill(start_color="76A4D8", end_color="76A4D8", fill_type="solid")

    for cell in ws['1:1']:
        cell.font = font_style  # 设置第一行的字体样式
        cell.fill = fill_color  # 设置第一行的背景颜色

    # 设置对齐方式
    alignment_center = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment_center

    # 设置边框样式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # 设置列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = (max_length * 1.6)
        ws.column_dimensions[column].width = adjusted_width

    # 设置条件格式规则
    optimal_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")

    formula_expression = f'{acc_col_index}1 = MAX(${acc_col_index}$1:${acc_col_index}$1000)'
    rule_green = FormulaRule(formula=[formula_expression], stopIfTrue=True, fill=optimal_fill)

    # 应用条件格式规则到对应单元格所在的一整列
    for index in range(ws.max_column):
        column_letter = get_column_letter(index+1)
        ws.conditional_formatting.add(f'{column_letter}1:{column_letter}1000', rule_green)

    # 保存文件
    wb.save(excel_path)


if __name__ == '__main__':
    # preprocess_data_and_config(r'./data/文本分类练习.csv')
    pretty_excel(r'./reports/training_report_20240801234335.xlsx')
